"""Amorçage : verse tes exports existants dans le format du dépôt.

À lancer UNE FOIS en local, puis commiter le résultat.

    python scripts/bootstrap_import.py ~/Downloads/mon_export_spotify

Le dossier doit contenir :
  - Streaming_History_Audio_*.json   (historique étendu)
  - YourLibrary.json                 (facultatif : likes, albums, artistes suivis)
  - Concerts*.xlsx                   (facultatif : ton fichier de concerts)
"""
import json, pathlib, sys, glob
from common import DATA, append_jsonl, write_json


def import_history(folder):
    files = sorted(glob.glob(str(pathlib.Path(folder) / "Streaming_History_Audio_*.json")))
    if not files:
        print("  aucun Streaming_History_Audio_*.json trouvé")
        return 0
    by_month, total = {}, 0
    for f in files:
        for r in json.load(open(f, encoding="utf-8")):
            if not r.get("master_metadata_track_name"):
                continue
            ts = r["ts"]
            by_month.setdefault(ts[:7], []).append({
                "ts": ts,
                "track": r["master_metadata_track_name"],
                "artist": r.get("master_metadata_album_artist_name", ""),
                "artists": [r.get("master_metadata_album_artist_name", "")],
                "album": r.get("master_metadata_album_album_name", ""),
                "ms": r.get("ms_played", 0),
                "uri": r.get("spotify_track_uri", "") or "",
                "album_uri": "",
                "src": "spotify-export",
            })
            total += 1
    for month, rows in sorted(by_month.items()):
        rows.sort(key=lambda x: x["ts"])
        p = DATA / "listens" / f"{month}.jsonl"
        if p.exists():
            p.unlink()
        append_jsonl(p, rows)
    newest = max(max(r["ts"] for r in rows) for rows in by_month.values())
    write_json(DATA / "state.json", {"last_played_at": newest, "bootstrap": True})
    print(f"  {total} écoutes réparties sur {len(by_month)} mois — dernière : {newest}")
    return total


def import_library(folder):
    hits = list(pathlib.Path(folder).glob("*YourLibrary.json"))
    if not hits:
        print("  pas de YourLibrary.json (facultatif)")
        return
    L = json.load(open(hits[0], encoding="utf-8"))
    write_json(DATA / "library.json", {
        "tracks": [{"track": t.get("track", ""), "artist": t.get("artist", ""),
                    "album": t.get("album", ""), "uri": t.get("uri", ""), "added": ""}
                   for t in L.get("tracks", [])],
        "albums": [{"album": a.get("album", ""), "artist": a.get("artist", ""),
                    "uri": a.get("uri", ""), "added": ""} for a in L.get("albums", [])],
        "artists": [{"artist": a.get("name", ""), "uri": a.get("uri", "")}
                    for a in L.get("artists", [])]})


def import_concerts(folder):
    hits = list(pathlib.Path(folder).glob("Concerts*.xlsx"))
    if not hits:
        print("  pas de Concerts*.xlsx (facultatif) — tu pourras saisir depuis le site")
        return
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("  openpyxl absent : pip install openpyxl, puis relance")
        return
    ws = load_workbook(hits[0], read_only=True).active
    rows = list(ws.iter_rows(values_only=True))
    hdr = [str(h or "").strip() for h in rows[0]]

    def col(*names):
        for n in names:
            if n in hdr:
                return hdr.index(n)
        return None
    ia, ig = col("Artiste"), col("Genre")
    iv, ic = col("Salle / Festival", "Salle"), col("Ville")
    it, idt = col("Type"), col("Date precise", "Date précise", "Date")
    ilk = col("Lien Setlist (setlist.fm)", "Setlist")
    out = []
    for r in rows[1:]:
        if not r or ia is None or not r[ia]:
            continue
        d = r[idt] if idt is not None else None
        out.append({"artist": str(r[ia]),
                    "genre": str(r[ig]) if ig is not None and r[ig] else "",
                    "venue": str(r[iv]) if iv is not None and r[iv] else "",
                    "city": str(r[ic]) if ic is not None and r[ic] else "",
                    "type": str(r[it]) if it is not None and r[it] else "",
                    "date": d.date().isoformat() if hasattr(d, "date") else str(d or "")[:10],
                    "setlist": str(r[ilk]) if ilk is not None and r[ilk] else ""})
    write_json(DATA / "concerts.json", sorted(out, key=lambda x: x["date"]))
    print(f"  {len(out)} concerts importés")


def main():
    if len(sys.argv) < 2:
        sys.exit(__doc__)
    folder = sys.argv[1]
    if not pathlib.Path(folder).is_dir():
        sys.exit(f"dossier introuvable : {folder}")
    print("Historique :"); import_history(folder)
    print("Bibliothèque :"); import_library(folder)
    print("Concerts :"); import_concerts(folder)
    print("\nTerminé. Lance maintenant : python scripts/build_site.py")


if __name__ == "__main__":
    main()
